import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_sales_excel(orders):
    """
    Generates a formatted Excel sheet containing a listing of sales orders.
    Returns a bytes buffer.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Orders Report"
    
    # Styles
    title_font = Font(name='Calibri', size=16, bold=True, color='2C3E50')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    bold_font = Font(name='Calibri', size=11, bold=True)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='E9ECEF'),
        right=Side(style='thin', color='E9ECEF'),
        top=Side(style='thin', color='E9ECEF'),
        bottom=Side(style='thin', color='E9ECEF')
    )
    
    # Title Block
    ws.merge_cells('A1:H1')
    ws['A1'] = "Sales Orders Summary Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = align_left
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = [
        "Order #", "Customer", "Company", "Date", 
        "Subtotal", "GST", "Discount", "Order Total", "Status"
    ]
    
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Data Rows
    row_idx = 4
    total_sales = 0.0
    for order in orders:
        ws.row_dimensions[row_idx].height = 20
        
        c1 = ws.cell(row=row_idx, column=1, value=order.order_number)
        c1.alignment = align_center
        c1.border = thin_border
        
        c2 = ws.cell(row=row_idx, column=2, value=order.customer.name)
        c2.alignment = align_left
        c2.border = thin_border
        
        c3 = ws.cell(row=row_idx, column=3, value=order.customer.company or "N/A")
        c3.alignment = align_left
        c3.border = thin_border
        
        c4 = ws.cell(row=row_idx, column=4, value=order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else "")
        c4.alignment = align_center
        c4.border = thin_border
        
        c5 = ws.cell(row=row_idx, column=5, value=float(order.subtotal))
        c5.number_format = '$#,##0.00'
        c5.alignment = align_right
        c5.border = thin_border
        
        c6 = ws.cell(row=row_idx, column=6, value=float(order.gst_amount))
        c6.number_format = '$#,##0.00'
        c6.alignment = align_right
        c6.border = thin_border
        
        c7 = ws.cell(row=row_idx, column=7, value=float(order.discount_amount))
        c7.number_format = '$#,##0.00'
        c7.alignment = align_right
        c7.border = thin_border
        
        c8 = ws.cell(row=row_idx, column=8, value=float(order.order_total))
        c8.number_format = '$#,##0.00'
        c8.alignment = align_right
        c8.border = thin_border
        
        c9 = ws.cell(row=row_idx, column=9, value=order.status)
        c9.alignment = align_center
        c9.border = thin_border
        
        total_sales += float(order.order_total)
        row_idx += 1
        
    # Summary Row
    ws.row_dimensions[row_idx].height = 22
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    summary_label = ws.cell(row=row_idx, column=1, value="Total Sales Value:")
    summary_label.font = bold_font
    summary_label.alignment = align_right
    summary_label.border = thin_border
    
    for col in range(1, 8):
        ws.cell(row=row_idx, column=col).border = thin_border
        
    summary_val = ws.cell(row=row_idx, column=8, value=total_sales)
    summary_val.font = bold_font
    summary_val.number_format = '$#,##0.00'
    summary_val.alignment = align_right
    summary_val.border = thin_border
    
    ws.cell(row=row_idx, column=9).border = thin_border
    
    # Auto-adjust columns widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid using merged cell lengths as column widths
            if cell.coordinate in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_inventory_excel(products):
    """
    Generates a formatted Excel sheet containing inventory stock status.
    Returns a bytes buffer.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    title_font = Font(name='Calibri', size=16, bold=True, color='2C3E50')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='00A09D', end_color='00A09D', fill_type='solid') # Teal style for inventory
    bold_font = Font(name='Calibri', size=11, bold=True)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='E9ECEF'),
        right=Side(style='thin', color='E9ECEF'),
        top=Side(style='thin', color='E9ECEF'),
        bottom=Side(style='thin', color='E9ECEF')
    )
    
    # Title Block
    ws.merge_cells('A1:G1')
    ws['A1'] = "Inventory Stock & Valuation Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = align_left
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = [
        "Product Name", "Product Code", "SKU", "Category", 
        "Cost Price", "Selling Price", "Stock Quantity", "Total Valuation"
    ]
    
    ws.row_dimensions[3].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Data Rows
    row_idx = 4
    total_val = 0.0
    for prod in products:
        ws.row_dimensions[row_idx].height = 20
        
        c1 = ws.cell(row=row_idx, column=1, value=prod.name)
        c1.alignment = align_left
        c1.border = thin_border
        
        c2 = ws.cell(row=row_idx, column=2, value=prod.code)
        c2.alignment = align_center
        c2.border = thin_border
        
        c3 = ws.cell(row=row_idx, column=3, value=prod.sku)
        c3.alignment = align_center
        c3.border = thin_border
        
        c4 = ws.cell(row=row_idx, column=4, value=prod.category.name if prod.category else "N/A")
        c4.alignment = align_left
        c4.border = thin_border
        
        c5 = ws.cell(row=row_idx, column=5, value=float(prod.cost_price))
        c5.number_format = '$#,##0.00'
        c5.alignment = align_right
        c5.border = thin_border
        
        c6 = ws.cell(row=row_idx, column=6, value=float(prod.selling_price))
        c6.number_format = '$#,##0.00'
        c6.alignment = align_right
        c6.border = thin_border
        
        c7 = ws.cell(row=row_idx, column=7, value=prod.stock_quantity)
        c7.number_format = '#,##0'
        c7.alignment = align_center
        c7.border = thin_border
        
        val = float(prod.cost_price) * prod.stock_quantity
        c8 = ws.cell(row=row_idx, column=8, value=val)
        c8.number_format = '$#,##0.00'
        c8.alignment = align_right
        c8.border = thin_border
        
        total_val += val
        row_idx += 1
        
    # Summary Row
    ws.row_dimensions[row_idx].height = 22
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    summary_label = ws.cell(row=row_idx, column=1, value="Total Inventory Asset Valuation:")
    summary_label.font = bold_font
    summary_label.alignment = align_right
    summary_label.border = thin_border
    
    for col in range(1, 8):
        ws.cell(row=row_idx, column=col).border = thin_border
        
    summary_val = ws.cell(row=row_idx, column=8, value=total_val)
    summary_val.font = bold_font
    summary_val.number_format = '$#,##0.00'
    summary_val.alignment = align_right
    summary_val.border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
